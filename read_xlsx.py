from numpy.core.records import array
import openpyxl
import pandas as pd
from pathlib import Path
xlsx_file = Path('./', 'covid19.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 
import numpy as np

sheet = wb_obj["data_student_learning_habit_cov"]
import matplotlib.pyplot as plt
import scipy as sp
import scipy.stats
import seaborn as sns


data_read_dict={
"Gender":[],
"Class":[],
"Grade_level":[],
"school_type":[],
"Sib":[],
"fa_job":[],
"mo_job":[],
"income":[],
"exam":[],
"Self_evaluation":[],
"English":[],
"Lh_before_Cov":[],
"Lh_in_Cov":[],
"Total_Online":[],
"LHInstruction":[],
"LH_w_Instruction":[],
"onl_instr":[],
"onl_noninstr":[],
"Total_offline":[],
"off_instr":[],
"off_noninstr":[],

"nec_prog":[],
"nec_habit":[],
"nec_teacher":[],
"nec_parent":[],
"nec_sib":[],
"nec_friend":[],

"eff_moti":[],
"eff_con":[],
"eff_supp":[],
"eff_env":[],
"eff_obj":[],
"eff_resource":[],
"eff_friend":[],

"kno_med":[],
"kno_covid":[],
"kno_env":[],
"kno_soci":[],
"kno_elearn":[]
}


def mean_confidence_interval(data, confidence=0.95):
    a = 1.0*np.array(data)
    n = len(a)
    m, se = np.mean(a), scipy.stats.sem(a)
    h = se * sp.stats.t._ppf((1+confidence)/2., n-1)
    return np.round(m-h,2), np.round(m+h,2)

for column in sheet.iter_cols(2, sheet.max_column) :
    
    column_data=[]
    for j in range(1,sheet.max_row):
        column_data.append(column[j].value)
    data_read_dict[column[0].value]=(column_data)  

def feature_group_by_(SORTED_BY,TYPE_NAMES):
    
    Total="Total"
    Total_Online="Total_Online"
    Total_Online="Total_Online"
    Total_offline="Total_offline"
    LHInstruction="LHInstruction"
    LH_w_Instruction="LH_w_Instruction"
    
    detail_Total="Total"
    detail_Total_Online="Total Online"
    detail_Total_offline="Total Offline"
    detail_LHInstruction="Hours with instruction"
    detail_LH_w_Instruction="Hours without instruction"
    
    
    feature_type=[Total,Total_Online,Total_offline,LHInstruction,LH_w_Instruction]
    feature_type_detail=[detail_Total,detail_Total_Online,detail_Total_offline,detail_LHInstruction,detail_LH_w_Instruction]
    group=[]

    
    for i in range(len(TYPE_NAMES)):
        group.append({feature_type[0]:[],feature_type[1]:[],feature_type[2]:[],feature_type[3]:[],feature_type[4]:[]})

    for i in range(0,len(data_read_dict[SORTED_BY])):
        _index=data_read_dict[SORTED_BY][i]
        group[_index-1][feature_type[1]].append(data_read_dict[Total_Online][i])
        group[_index-1][feature_type[2]].append(data_read_dict[Total_offline][i])
        group[_index-1][feature_type[3]].append(data_read_dict[LHInstruction][i])
        group[_index-1][feature_type[4]].append(data_read_dict[LH_w_Instruction][i])


    d_Total="Total"
    d_Mean="Mean"
    d_Std_Deviation="Std.Deviation"
    d_Lower_Bound="Lower Bound"
    d_Upper_Bound="Upper Bound"
    d_Min="Min"
    d_Max="Max"

    column_name=[SORTED_BY,"Learning hours",d_Total,d_Mean,d_Std_Deviation,d_Lower_Bound,d_Upper_Bound,d_Min,d_Max]    
    columns_data=[]
    columns_data.append([])
    for _ in range(len(column_name)):
        columns_data.append([])
    
    for i in range(len(TYPE_NAMES)):
        _total=len(group[i][feature_type[1]])
        for j in range(1,len(feature_type)):
                mean_feature=np.round(np.mean(group[i][feature_type[j]]),2)
                std_feature=np.round(np.std(group[i][feature_type[j]]),3)
                max_feature=np.round(np.max(group[i][feature_type[j]]),2)
                min_feature=np.round(np.min(group[i][feature_type[j]]),2)

                (upper,lower)=mean_confidence_interval(group[i][feature_type[j]])
                k=0
                columns_data[k].append(TYPE_NAMES[i])
                k=k+1
                columns_data[k].append(feature_type_detail[j])
                k=k+1
                columns_data[k].append(_total)
                k=k+1
                columns_data[k].append(mean_feature)
                k=k+1
                columns_data[k].append(std_feature)
                k=k+1
                columns_data[k].append(upper)
                k=k+1
                columns_data[k].append(lower)
                k=k+1
                columns_data[k].append(min_feature)
                k=k+1
                columns_data[k].append(max_feature)
           
    
    data = {column_name[0]: columns_data[0],
            column_name[1]: columns_data[1],
            column_name[2]: columns_data[2],
            column_name[3]: columns_data[3],
            column_name[4]: columns_data[4],
            column_name[5]: columns_data[5],
            column_name[6]: columns_data[6],
            column_name[7]: columns_data[7],
            column_name[8]: columns_data[8],
            
            }

    df = pd.DataFrame(data, columns=column_name)
    
    return df
            

#feature_group_by_("Gender",["Male","Female","Not public"])

#feature_group_by_("school_type",["Public school (normal)",	"Public school (Gifted)",	"Private school (normal)",	"International school"])

#feature_group_by_("Grade_level",["Secondary school","High school"])

# feature_group_by_("income",["Less than 430 USD"	,
# "From 430 to under 860 USD",
# "From 860 to under 1,290 USD",
# "From 1,290 to under 1,720 USD",
# "From 1,720 to under 2,150 USD"	,
# "More than 2,150 USD"])

#feature_group_by_("Sib",["One",	"Two"	,"Three"	,"Over four"])

#feature_group_by_("fa_job",["Stem"	,"Social siences"	,"Free"	,"Others"])

#feature_group_by_("mo_job",["Stem"	,"Social siences"	,"Free"	,"Others"])

#feature_group_by_("exam",["A (Mathematics, Physics, Chemistry)",	"A1 (Mathematics, Physics, English)",	"B (Mathematics, Biology, Chemistry)","C (Literature, History, Geography)",	"D (Literature, Foreign Language, Mathematics)"	,"Others"])

#feature_group_by_("Lh_before_Cov",["under 4h"	,"from 4 to 7h"	,"over 7h"])

#feature_group_by_("Lh_in_Cov",["under 4h"	,"from 4 to 7h"	,"over 7h"])

#feature_group_by_("English",["Below Average",	"Average","Good",	"Excellence"])

#Internet and Additional Resources
#feature_group_by_("eff_resource",["Strongly disagree"	,"Disagree"	,"Neither agree nor disagree",	"Agree"	,"Strongly agree"])


	
    	
# Draw Plot
# plt.figure(figsize=(16,10), dpi= 80)

# _data={"val":[1,2,1,1,11,1,4,5,6,9],"name":"dinh"}
# sns.kdeplot( data=_data, shade=True, color="g", label="Cyl=4", alpha=.7)
# #sns.kdeplot(df.loc[df['cyl'] == 4, "cty"], shade=True, color="g", label="Cyl=4", alpha=.7)
# # sns.kdeplot(df.loc[df['cyl'] == 5, "cty"], shade=True, color="deeppink", label="Cyl=5", alpha=.7)
# # sns.kdeplot(df.loc[df['cyl'] == 6, "cty"], shade=True, color="dodgerblue", label="Cyl=6", alpha=.7)
# # sns.kdeplot(df.loc[df['cyl'] == 8, "cty"], shade=True, color="orange", label="Cyl=8", alpha=.7)

# # Decoration
# plt.title('Density Plot of City Mileage by n_Cylinders', fontsize=22)
# plt.legend()
# plt.show()